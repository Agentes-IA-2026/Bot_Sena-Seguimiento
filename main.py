import os
import re
import glob
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from datetime import datetime

# ── SSL: configurar CA bundle ANTES de cualquier conexión HTTPS ───────────────
# ssl_config.py lee SSL_CA_BUNDLE / REQUESTS_CA_BUNDLE / certifi y setea las
# variables de entorno que usan urllib3, google-auth y httplib2.
from ssl_config import configure_environment
configure_environment()

from Core.drive_manager import (
    conectar_drive,
    verificar_evidencias_en_carpeta,
)
from Core.document_analyzer import extraer_nombres_evidencias_manual, extraer_nombres_evidencias

# Fase 2: motor enriquecido
from bot.auditor import MotorAuditoria, EstadoAuditoria, es_entregada
from bot.drive_adapter import listar_archivos_con_info_service
from sincronizador_tabla import cargar_actividades

# ── SUPABASE ──────────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from supabase_client import get_supabase
DOCENTE_ID = os.environ.get("DOCENTE_ID", "d114d85c-fb7e-4a79-a446-fb90d652eec2")

# ── CONFIGURACIÓN ─────────────────────────────────────────────────────────────
CARPETA_LISTADOS = "assets/Listados"
CARPETA_GUIAS    = "assets/Guias_Referencia"
CARPETA_REPORTES = "logs/reportes"

MAPA_PROGRAMAS = {
    "asesoria comercial":       "Asistencia_comercial",
    "asistencia comercial":     "Asistencia_comercial",
    "comunicacion comercial":   "Comunicacion_y_marketing",
    "comercial y marketing":    "Comunicacion_y_marketing",
    "comunicacion y marketing": "Comunicacion_y_marketing",
    "comunicacion":             "Comunicacion_y_marketing",
    "marketing":                "Comunicacion_y_marketing",
    "ventas de productos":      "Ventas_de_productos_en_linea",
    "productos en linea":       "Ventas_de_productos_en_linea",
    "ventas":                   "Ventas_de_productos_en_linea",
}

# Mapeo directo ficha → programa canónico (fuente de verdad para la resolución).
# Tiene prioridad sobre el texto en la celda C3 del Excel de listado.
FICHA_A_PROGRAMA: dict[str, str] = {
    "3414936": "Asistencia_comercial",
    "3414937": "Asistencia_comercial",
    "3415417": "Comunicacion_y_marketing",
    "3415087": "Comunicacion_y_marketing",
    "3458742": "Ventas_de_productos_en_linea",
}

# Guía transversal: se incluye en todos los programas sin importar el mapeo.
GUIA_INDUCCION = "GUIA_00_INDUCCION"

# Subcarpetas esperadas por guía dentro del portafolio del aprendiz.
# Si la subcarpeta existe en Drive, la auditoría de esa guía se limita a ella.
# Si no existe, se usa el portafolio completo como fallback.
SUBCARPETA_GUIA: dict[str, str] = {
    "GUIA_01": "analisis",
}

# Mapeo canonical → palabras clave que aparecen en actividades_parametrizadas.programa.
# Necesario porque el texto guardado en la tabla usa el nombre largo del programa
# ("Técnico En Asesoría Comercial") en lugar del canonical corto ("Asistencia_comercial").
# Las palabras clave se comparan contra el valor normalizado (sin tildes, minúsculas).
CANONICAL_A_PALABRAS_CLAVE: dict[str, list[str]] = {
    "Asistencia_comercial"        : ["asesoria comercial"],
    "Comunicacion_y_marketing"    : ["comunicacion comercial", "marketing"],
    "Ventas_de_productos_en_linea": ["ventas de productos"],
}

COLOR_VERDE    = "C6EFCE"   # OK
COLOR_ROJO     = "FFC7CE"   # FALTA
COLOR_AMARILLO = "FFEB9C"   # sin link / NOMBRE_DIFERENTE / CARPETA_INCORRECTA
COLOR_NARANJA  = "F4B942"   # FORMATO_INCORRECTO
COLOR_AZUL_CLR = "BDD7EE"   # EVIDENCIA_CRUZADA
COLOR_GRIS_CLR = "EFEFEF"   # OPCIONAL_FALTANTE
COLOR_GRIS     = "D9D9D9"   # encabezados


# ── SUPABASE: GUARDAR ─────────────────────────────────────────────────────────

def _guardar_en_supabase(resultados: list[dict], ficha: str, colegio: str, siglas: str):
    """
    Guarda resultados en Supabase de forma ACUMULATIVA:
    - Borra solo las guías que se acaban de auditar (no toda la ficha).
    - Las guías auditadas en sesiones anteriores se conservan.
    - Re-auditar la misma guía actualiza sus registros sin duplicar.
    """
    try:
        db    = get_supabase()
        fecha = datetime.now().isoformat()

        registros = []
        for aprendiz in resultados:
            for nombre_guia, evidencias in aprendiz.get("guias", {}).items():
                guia_limpia = re.sub(r'\.pdf$', '', nombre_guia, flags=re.IGNORECASE)
                guia_limpia = guia_limpia.replace('_', ' ')
                guia_limpia = re.sub(r' {2,}', ' ', guia_limpia).strip()

                for evidencia, entregado in evidencias.items():
                    ev_limpia = re.sub(r'\.\w+$', '', evidencia).replace('_', ' ').strip()
                    registros.append({
                        "docente_id" : DOCENTE_ID,
                        "ficha"      : ficha,
                        "colegio"    : f"{siglas} — {colegio}",
                        "estudiante" : aprendiz["nombre"],
                        "guia"       : guia_limpia,
                        "evidencia"  : ev_limpia,
                        "entregado"  : bool(entregado),
                        "fecha"      : fecha,
                    })

        if not registros:
            print("   ⚠️  Sin registros para guardar")
            return

        # ── DELETE acotado: solo las guías auditadas en esta sesión ──────────
        # Preserva registros de guías anteriores (comportamiento acumulativo).
        guias_sesion = sorted({r["guia"] for r in registros})
        for guia in guias_sesion:
            db.table("verificaciones")\
              .delete()\
              .eq("docente_id", DOCENTE_ID)\
              .eq("ficha", ficha)\
              .eq("guia", guia)\
              .execute()
        print(f"   🗑️  Guías actualizadas: {guias_sesion}")

        # ── INSERT en lotes de 500 ────────────────────────────────────────────
        for i in range(0, len(registros), 500):
            db.table("verificaciones").insert(registros[i:i+500]).execute()

        print(f"   ☁️  {len(registros)} registros → tabla 'verificaciones' ✅")

    except Exception as e:
        print(f"   ❌ Error guardando en tabla 'verificaciones': {e}")


# ── HELPERS ───────────────────────────────────────────────────────────────────

def _leer_celda_texto(celda) -> str:
    """Lee una celda de Excel y la convierte siempre a texto limpio."""
    val = celda.value
    if val is None:
        return ""
    # Si es número (int o float) convertir directamente
    if isinstance(val, (int, float)):
        return str(int(val)).strip()
    return str(val).strip()


def _normalizar(texto: str) -> str:
    texto = texto.lower()
    for a, b in [('á','a'),('é','e'),('í','i'),('ó','o'),('ú','u'),('ñ','n')]:
        texto = texto.replace(a, b)
    return texto


def _resolver_carpeta_guias(nombre_programa: str) -> str | None:
    prog_norm = _normalizar(nombre_programa)
    for clave, carpeta in MAPA_PROGRAMAS.items():
        if clave in prog_norm:
            ruta = os.path.join(CARPETA_GUIAS, carpeta)
            return ruta if os.path.isdir(ruta) else None
    return None


def obtener_programa_por_ficha(ficha: str) -> str | None:
    """Lookup local en FICHA_A_PROGRAMA (fallback cuando Supabase no responde)."""
    return FICHA_A_PROGRAMA.get(str(ficha).strip())


def obtener_programa_desde_supabase(ficha: str) -> str | None:
    """
    Consulta ficha_programa ⟶ programas.canonical en Supabase.
    Retorna el canonical del programa o None si no hay registro activo.
    """
    try:
        db   = get_supabase()
        resp = (
            db.table("ficha_programa")
            .select("programas(canonical)")
            .eq("ficha", str(ficha).strip())
            .eq("activa", True)
            .limit(1)
            .execute()
        )
        data = resp.data
        if data and data[0].get("programas"):
            return data[0]["programas"].get("canonical")
        return None
    except Exception as e:
        print(f"   ⚠️  Error consultando ficha_programa: {e}")
        return None


def resolver_programa_ficha(ficha: str) -> tuple[str | None, str]:
    """
    Resuelve el programa canónico para una ficha.
    Orden de prioridad:
      1. Supabase (ficha_programa + programas)
      2. Fallback local (FICHA_A_PROGRAMA)

    Retorna (canonical, fuente)
    fuente: "supabase" | "fallback_local" | "no_encontrado"
    """
    print(f"   🔎 Buscando programa de ficha en Supabase: {ficha}")

    canonical = obtener_programa_desde_supabase(ficha)
    if canonical:
        print(f"   ✅ Programa resuelto desde Supabase: {canonical}")
        return canonical, "supabase"

    fallback = FICHA_A_PROGRAMA.get(str(ficha).strip())
    if fallback:
        print(f"   ⚠️  Fallback local usado: {fallback}")
        return fallback, "fallback_local"

    print(f"   ❌ No se encontró programa para la ficha: {ficha}")
    return None, "no_encontrado"


def obtener_guias_por_programa(programa: str) -> list[str]:
    """
    Retorna la lista ordenada de IDs de guía para el programa dado.
    GUIA_00_INDUCCION se garantiza siempre (es transversal a todos los programas).

    Consulta actividades_parametrizadas en Supabase:
      - todas las guías del programa indicado,
      - más GUIA_INDUCCION si no estuviera ya incluida.
    """
    acts_programa  = cargar_actividades(programa=programa)
    acts_induccion = cargar_actividades(guia=GUIA_INDUCCION)

    guias = sorted({
        a["guia"]
        for a in acts_programa + acts_induccion
        if a.get("guia")
    })
    return guias


def _extraer_id_carpeta(link: str) -> str | None:
    if not link or not isinstance(link, str):
        return None
    link = link.strip()
    for patron in [
        r'/folders/([a-zA-Z0-9_-]+)',
        r'/file/d/([a-zA-Z0-9_-]+)',
        r'[?&]id=([a-zA-Z0-9_-]+)',
        r'open\?id=([a-zA-Z0-9_-]+)',
    ]:
        m = re.search(patron, link)
        if m:
            return m.group(1)
    if "http" not in link and len(link) > 10:
        return link
    return None


def _leer_evidencias_de_guias(carpeta_guias: str) -> dict:
    guias = {}
    todos = []
    if os.path.isdir(carpeta_guias):
        for archivo in os.listdir(carpeta_guias):
            if archivo.lower().endswith('.pdf'):
                todos.append(os.path.join(carpeta_guias, archivo))
    todos = sorted(todos)

    if not todos:
        print(f"❌ No se encontraron guías PDF en: {carpeta_guias}")
        return guias

    for pdf in todos:
        nombre = os.path.basename(pdf)
        # Primero usar lista manual (limpia y confiable)
        evidencias = extraer_nombres_evidencias_manual(pdf)
        if not evidencias:
            # Solo si no hay lista manual, extraer automáticamente del PDF
            evidencias = extraer_nombres_evidencias(pdf, debug=False)
        if evidencias:
            guias[nombre] = evidencias
            print(f"   📋 {nombre}: {len(evidencias)} evidencia(s)")
        else:
            # Incluir la guía con marcador para que aparezca en el dashboard
            guias[nombre] = ["Evidencia pendiente de configurar"]
            print(f"   ⚠️  Sin evidencias configuradas para: {nombre}")
    return guias


def _leer_aprendices(hoja) -> list[dict]:
    aprendices = []
    for fila in hoja.iter_rows(min_row=10, max_row=200):
        nombre_val = fila[1].value
        celda_link = fila[10]

        nombre = str(nombre_val).strip() if nombre_val else None
        if not nombre or nombre in ["Nombre", "TI", "PPT", "CC", "None", "nan", ""]:
            continue

        if celda_link.hyperlink:
            link = celda_link.hyperlink.target
        elif celda_link.value:
            link = str(celda_link.value).strip()
        else:
            link = None

        aprendices.append({"nombre": nombre, "link": link})
    return aprendices


# ── REPORTE EXCEL LOCAL ───────────────────────────────────────────────────────

def _generar_reporte_excel(ficha: str, programa: str, resultados: list[dict],
                           guias: dict, ruta_salida: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    for nombre_guia, evidencias in guias.items():
        nombre_hoja = re.sub(r'\.pdf$', '', nombre_guia, flags=re.IGNORECASE)[:28]
        ws = wb.create_sheet(title=nombre_hoja)

        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=2 + len(evidencias))
        ct = ws.cell(row=1, column=1)
        ct.value = f"REPORTE FICHA {ficha} | {programa} | {nombre_guia}"
        ct.font = Font(bold=True, size=11, color="FFFFFF")
        ct.fill = PatternFill("solid", fgColor="2E4057")
        ct.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=2 + len(evidencias))
        cf = ws.cell(row=2, column=1)
        cf.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        cf.font = Font(italic=True, size=9)
        cf.alignment = Alignment(horizontal="right")

        encabezados = ["Aprendiz"] + [
            re.sub(r'\.\w+$', '', ev) for ev in evidencias
        ] + ["Total", "%"]

        for col, enc in enumerate(encabezados, 1):
            c = ws.cell(row=3, column=col, value=enc)
            c.font = Font(bold=True, size=9)
            c.fill = PatternFill("solid", fgColor=COLOR_GRIS)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = borde
        ws.row_dimensions[3].height = 40

        for fila_idx, aprendiz in enumerate(resultados, 4):
            reporte_guia = aprendiz.get("guias", {}).get(nombre_guia, {})
            sin_link = not aprendiz.get("link")

            c_nombre = ws.cell(row=fila_idx, column=1, value=aprendiz["nombre"])
            c_nombre.font = Font(size=9)
            c_nombre.border = borde
            c_nombre.alignment = Alignment(vertical="center")

            entregadas = 0
            for col_idx, ev in enumerate(evidencias, 2):
                c = ws.cell(row=fila_idx, column=col_idx)
                if sin_link:
                    c.value = "—"
                    c.fill = PatternFill("solid", fgColor=COLOR_AMARILLO)
                elif reporte_guia.get(ev):
                    c.value = "✓"
                    c.fill = PatternFill("solid", fgColor=COLOR_VERDE)
                    entregadas += 1
                else:
                    c.value = "✗"
                    c.fill = PatternFill("solid", fgColor=COLOR_ROJO)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(size=9, bold=True)
                c.border = borde

            total_ev = len(evidencias)
            pct = (entregadas / total_ev * 100) if total_ev and not sin_link else 0

            c_total = ws.cell(row=fila_idx, column=2 + len(evidencias),
                              value=f"{entregadas}/{total_ev}" if not sin_link else "—")
            c_pct   = ws.cell(row=fila_idx, column=3 + len(evidencias),
                              value=f"{pct:.0f}%" if not sin_link else "—")

            for c in [c_total, c_pct]:
                c.font = Font(size=9, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = borde
                if not sin_link:
                    color = COLOR_VERDE if pct == 100 else (COLOR_ROJO if pct < 50 else COLOR_AMARILLO)
                    c.fill = PatternFill("solid", fgColor=color)

        ws.column_dimensions["A"].width = 30
        for col in range(2, 2 + len(evidencias)):
            letra = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[letra].width = 14
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + len(evidencias))].width = 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(3 + len(evidencias))].width = 8
        ws.freeze_panes = "B4"

    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)
    print(f"   💾 Reporte Excel: {ruta_salida}")


# ── REPORTE EXCEL v2 (FASE 2) ────────────────────────────────────────────────

# Mapeo estado → (símbolo, color_hex)
_ESTADO_ESTILO: dict[str, tuple[str, str]] = {
    "OK"                 : ("✓", COLOR_VERDE),
    "NOMBRE_DIFERENTE"   : ("~", COLOR_AMARILLO),
    "CARPETA_INCORRECTA" : ("~", COLOR_AMARILLO),
    "FORMATO_INCORRECTO" : ("⚠", COLOR_NARANJA),   # entregada, pero formato distinto
    "EVIDENCIA_CRUZADA"  : ("↔", COLOR_AZUL_CLR),
    "FALTA"              : ("✗", COLOR_ROJO),
    "OPCIONAL_FALTANTE"  : ("○", COLOR_GRIS_CLR),
}
_LEYENDA = [
    ("✓  OK — entregada con nombre correcto",          COLOR_VERDE),
    ("~  Encontrada, nombre o carpeta diferente",       COLOR_AMARILLO),
    ("⚠  Encontrada, formato diferente (cuenta como entregada)", COLOR_NARANJA),
    ("↔  Evidencia cruzada — revisar con el aprendiz", COLOR_AZUL_CLR),
    ("✗  Falta — no se encontró evidencia",            COLOR_ROJO),
    ("○  Opcional — no subida (no penaliza)",           COLOR_GRIS_CLR),
    ("—  Sin link de Drive",                            "FFEB9C"),
]


def _generar_reporte_excel_v2(
    ficha: str,
    programa: str,
    resultados: list[dict],
    ruta_salida: str,
    guias_a_incluir: list[str] | None = None,
):
    """
    Genera el reporte Excel. Fuente de verdad EXCLUSIVA: actividades_parametrizadas.

    Para cada guía consulta cargar_actividades(guia=nombre_guia) en tiempo de ejecución,
    de modo que los encabezados de columna provienen siempre de actividad_resumen
    en Supabase, nunca de nombres de archivos Drive, IDs internos ni LISTAS_MANUALES.

    Parámetros
    ----------
    resultados       — [{ nombre, link, guias_v2: {guia_id: {actividad_id: estado}} }]
    guias_a_incluir  — lista ordenada de IDs de guía a incluir en el reporte.
                       Si None, se deriva de la unión de guias_v2.keys() en resultados.
    """
    # ── Determinar qué guías incluir ──────────────────────────────────────────
    if not guias_a_incluir:
        guias_a_incluir = sorted({
            guia
            for r in resultados
            for guia in r.get("guias_v2", {}).keys()
        })

    if not guias_a_incluir:
        print("   ⚠️  Sin guías auditadas — reporte vacío.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    hojas_generadas = 0

    for nombre_guia in guias_a_incluir:

        # ── Fuente de verdad: consulta directa a actividades_parametrizadas ──
        # Nunca usamos grupos_v2 ni ningún caché aquí.
        actividades = sorted(
            [
                a for a in cargar_actividades(guia=nombre_guia)
                if a.get("actividad_id") and a.get("actividad_resumen")
            ],
            key=lambda a: a["actividad_id"],
        )

        if not actividades:
            print(f"   ⚠️  Sin actividades en actividades_parametrizadas para: {nombre_guia}")
            print(f"        Ejecuta: python sincronizador_tabla.py --guia {nombre_guia}")
            continue

        nombre_hoja = nombre_guia[:28]
        ws = wb.create_sheet(title=nombre_hoja)
        n_cols = len(actividades)
        hojas_generadas += 1

        # Denominador del % = solo obligatorias (OPCIONAL_FALTANTE no penaliza)
        n_obligatorias = sum(1 for a in actividades if a.get("obligatoria", True))
        denominador    = n_obligatorias if n_obligatorias > 0 else len(actividades)

        # ── Fila 1: título ───────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=2 + n_cols)
        ct = ws.cell(row=1, column=1)
        ct.value     = f"REPORTE FICHA {ficha} | {programa} | {nombre_guia}"
        ct.font      = Font(bold=True, size=11, color="FFFFFF")
        ct.fill      = PatternFill("solid", fgColor="2E4057")
        ct.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # ── Fila 2: metadatos ────────────────────────────────────────────────
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=2 + n_cols)
        cf = ws.cell(row=2, column=1)
        cf.value = (
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  "
            f"|  Actividades: {len(actividades)}  "
            f"|  Obligatorias: {denominador}  "
            f"|  Fuente: actividades_parametrizadas"
        )
        cf.font      = Font(italic=True, size=9)
        cf.alignment = Alignment(horizontal="right")

        # ── Fila 3: encabezados ──────────────────────────────────────────────
        # Texto de columna = actividad_resumen (orden fijo por actividad_id)
        encabezados = (
            ["Aprendiz"]
            + [a["actividad_resumen"] for a in actividades]
            + ["Total", "%"]
        )
        for col, enc in enumerate(encabezados, 1):
            c = ws.cell(row=3, column=col, value=enc)
            c.font      = Font(bold=True, size=9)
            c.fill      = PatternFill("solid", fgColor=COLOR_GRIS)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = borde
        ws.row_dimensions[3].height = 48

        # ── Filas de aprendices ──────────────────────────────────────────────
        for fila_idx, aprendiz in enumerate(resultados, 4):
            # reporte_guia  = { actividad_id: estado_str }
            # reporte_adv   = { actividad_id: texto_advertencia } — solo FORMATO_INCORRECTO
            reporte_guia = aprendiz.get("guias_v2",     {}).get(nombre_guia, {})
            reporte_adv  = aprendiz.get("guias_v2_adv", {}).get(nombre_guia, {})
            sin_link     = not aprendiz.get("link")

            c_nombre           = ws.cell(row=fila_idx, column=1, value=aprendiz["nombre"])
            c_nombre.font      = Font(size=9)
            c_nombre.border    = borde
            c_nombre.alignment = Alignment(vertical="center")

            entregadas_ok = 0
            for col_idx, act in enumerate(actividades, 2):
                aid    = act["actividad_id"]          # clave de cruce
                oblig  = act.get("obligatoria", True)
                estado = reporte_guia.get(aid)        # None → auditoría sin dato

                c = ws.cell(row=fila_idx, column=col_idx)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font      = Font(size=9, bold=True)
                c.border    = borde

                if sin_link:
                    c.value = "—"
                    c.fill  = PatternFill("solid", fgColor=COLOR_AMARILLO)
                elif estado is None:
                    c.value = "?"
                    c.fill  = PatternFill("solid", fgColor=COLOR_GRIS)
                else:
                    simbolo, color = _ESTADO_ESTILO.get(estado, ("✗", COLOR_ROJO))
                    c.value = simbolo
                    c.fill  = PatternFill("solid", fgColor=color)

                    # Formato incorrecto sigue siendo ENTREGADO — el tipo es solo advertencia.
                    # Se cuenta igual que OK/NOMBRE_DIFERENTE/CARPETA_INCORRECTA.
                    if oblig and es_entregada({"estado": estado}):
                        entregadas_ok += 1

                    # Añadir comentario de celda cuando el formato difiere del esperado
                    if estado == EstadoAuditoria.FORMATO_INCORRECTO.value:
                        adv_texto = reporte_adv.get(aid, "formato diferente al esperado")
                        c.comment = Comment(adv_texto, "GuíaBot")

            # ── Total y % ────────────────────────────────────────────────────
            pct = (entregadas_ok / denominador * 100) if denominador and not sin_link else 0
            c_total = ws.cell(
                row=fila_idx, column=2 + n_cols,
                value=f"{entregadas_ok}/{denominador}" if not sin_link else "—",
            )
            c_pct = ws.cell(
                row=fila_idx, column=3 + n_cols,
                value=f"{pct:.0f}%" if not sin_link else "—",
            )
            for c in [c_total, c_pct]:
                c.font      = Font(size=9, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = borde
                if not sin_link:
                    color_pct = (COLOR_VERDE if pct == 100
                                 else COLOR_ROJO if pct < 50
                                 else COLOR_AMARILLO)
                    c.fill = PatternFill("solid", fgColor=color_pct)

        # ── Leyenda ──────────────────────────────────────────────────────────
        fila_leyenda = len(resultados) + 5
        cl           = ws.cell(row=fila_leyenda, column=1, value="LEYENDA:")
        cl.font      = Font(bold=True, size=8)
        for i, (texto, color) in enumerate(_LEYENDA, 1):
            c           = ws.cell(row=fila_leyenda + i, column=1, value=texto)
            c.fill      = PatternFill("solid", fgColor=color)
            c.font      = Font(size=8)
            c.border    = borde
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[fila_leyenda].height = 14

        # ── Anchos de columna ────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 32
        for col in range(2, 2 + n_cols):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + n_cols)].width = 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(3 + n_cols)].width = 8
        ws.freeze_panes = "B4"

    if hojas_generadas == 0:
        print("   ⚠️  Reporte vacío: ninguna guía tiene actividades en actividades_parametrizadas.")
        print("        Ejecuta: python sincronizador_tabla.py para cargar el Excel maestro.")
        return

    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)
    print(f"   💾 Reporte Excel v2: {ruta_salida}  ({hojas_generadas} hoja(s))")


# ── FASE 2: CARGA DE ACTIVIDADES Y GUARDADO ENRIQUECIDO ──────────────────────

def _cargar_actividades_por_guia(
    programa_raw: str,
    programa_canonical: str | None = None,
) -> dict[str, list[dict]]:
    """
    Carga actividades desde Supabase y las agrupa por guía.
    Retorna { "GUIA_00_INDUCCION": [...], "GUIA_01_...": [...], ... }

    Estrategia de matching (en orden):
      1. Exact eq por programa_canonical  → más eficiente
      2. Filtro local por valor normalizado → tolerante a mayúsculas/tildes/underscores
      3. GUIA_INDUCCION se añade siempre aunque pertenezca a otro programa en Supabase.

    Si todo falla retorna {} y main.py usa el fallback de PDFs locales.
    """
    # ── Resolver nombre canónico del programa ─────────────────────────────────
    if not programa_canonical:
        prog_norm = _normalizar(programa_raw)
        for clave, carpeta in MAPA_PROGRAMAS.items():
            if _normalizar(clave) in prog_norm or prog_norm in _normalizar(clave):
                programa_canonical = carpeta
                break

    if not programa_canonical:
        print("   ⚠️  No se pudo resolver el programa canónico — sin guías cargadas")
        return {}

    prog_canon_norm = _normalizar(programa_canonical)
    print(f"\n   🔎 Buscando actividades para programa = '{programa_canonical}'")

    # ── Estrategia 1: match exacto ────────────────────────────────────────────
    actividades = cargar_actividades(programa=programa_canonical)
    print(f"   🔎 Match exacto     → {len(actividades)} filas")

    # ── Estrategia 2: fallback con carga total + filtro normalizado ───────────
    if not actividades:
        print("   🔎 Sin filas con match exacto — cargando universo completo...")
        todas = cargar_actividades()          # sin filtros
        print(f"   🔎 Total en tabla   → {len(todas)} filas")

        # Mostrar valores únicos del campo programa para diagnóstico
        progs_en_tabla = sorted({a.get("programa", "") for a in todas if a.get("programa")})
        print(f"   🔎 Programas en tabla: {progs_en_tabla}")

        # Filtro por normalización (tolera mayúsculas, tildes, espacios vs underscores)
        actividades = [
            a for a in todas
            if _normalizar(a.get("programa", "")) == prog_canon_norm
        ]
        print(f"   🔎 Filtro local     → {len(actividades)} filas")

        if not actividades:
            # Estrategia 3: containment normalizado
            actividades = [
                a for a in todas
                if prog_canon_norm in _normalizar(a.get("programa", ""))
                or _normalizar(a.get("programa", "")) in prog_canon_norm
            ]
            print(f"   🔎 Containment      → {len(actividades)} filas")

        if not actividades:
            # Estrategia 4: palabras clave del canonical (cubre "asistencia" vs "asesoria")
            palabras = CANONICAL_A_PALABRAS_CLAVE.get(programa_canonical, [])
            if palabras:
                actividades = [
                    a for a in todas
                    if any(kw in _normalizar(a.get("programa", "")) for kw in palabras)
                ]
                progs_alias = sorted({a.get("programa", "") for a in actividades})
                print(f"   🔎 Alias por keywords {palabras}")
                print(f"   🔎 Valores reales encontrados: {progs_alias}")
                print(f"   🔎 Estrategia 4     → {len(actividades)} filas")

    # ── Añadir GUIA_INDUCCION si no llegó con el programa ────────────────────
    guias_encontradas = {a.get("guia") for a in actividades}
    if GUIA_INDUCCION not in guias_encontradas:
        acts_00 = cargar_actividades(guia=GUIA_INDUCCION)
        actividades += acts_00
        print(f"   🔎 GUIA_00 añadida  → {len(acts_00)} filas adicionales")

    if not actividades:
        print("   ⚠️  Sin actividades tras todas las estrategias — se usará fallback PDF")
        return {}

    # ── Agrupar por guía (sin filtro adicional: todo lo que llegó es válido) ──
    guias_distintas = sorted({a.get("guia") for a in actividades if a.get("guia")})
    print(f"   🔎 Guías distintas  → {guias_distintas}")

    grupos: dict[str, list[dict]] = {}
    for act in actividades:
        g = act.get("guia")
        if g:
            grupos.setdefault(g, []).append(act)

    print(f"\n   📋 Actividades agrupadas por guía:")
    for guia, actos in sorted(grupos.items()):
        print(f"      • {guia}: {len(actos)} actividad(es)")

    return grupos


def _resolver_archivos_para_guia(
    archivos: list[dict],
    nombre_guia: str,
) -> list[dict]:
    """
    Devuelve el subconjunto de archivos Drive relevantes para una guía.

    Si SUBCARPETA_GUIA define una subcarpeta para esta guía (comparación normalizada),
    filtra a los archivos cuyo folder_path o carpeta_padre la contiene.
    Si la subcarpeta no existe entre los archivos, devuelve todos (fallback).
    Guías sin entrada en SUBCARPETA_GUIA reciben siempre la lista completa.
    """
    clave = next(
        (k for k in SUBCARPETA_GUIA if k in nombre_guia.upper()),
        None,
    )
    if not clave:
        return archivos

    sub_norm = _normalizar(SUBCARPETA_GUIA[clave])   # ej. "analisis"
    filtrados = [
        a for a in archivos
        if sub_norm in _normalizar(
            a.get("folder_path", "") + " " + a.get("carpeta_padre", "")
        )
    ]
    # Si la subcarpeta no existe en Drive, auditar con todos los archivos
    return filtrados if filtrados else archivos


def _guardar_auditoria_enriquecida(
    resultados_v2: list[dict],
    ficha: str,
    colegio: str,
    siglas: str,
    nombre_aprendiz: str,
):
    """
    Guarda resultados del motor enriquecido en la tabla evidencias_auditoria.
    No borra registros anteriores (append-only para historial).
    """
    if not resultados_v2:
        return
    try:
        db    = get_supabase()
        ahora = datetime.now().isoformat()

        registros = [
            {
                "docente_id"        : DOCENTE_ID,
                "ficha"             : ficha,
                "colegio"           : f"{siglas} — {colegio}",
                "estudiante"        : nombre_aprendiz,
                "actividad_id"      : r.get("actividad_id"),
                "guia"              : r.get("guia"),
                "programa"          : r.get("programa"),
                "actividad_resumen" : r.get("actividad_resumen"),
                "nombre_esperado"   : r.get("nombre_esperado"),
                "obligatoria"       : r.get("obligatoria", True),
                "estado"            : r.get("estado"),
                "archivo_encontrado": r.get("archivo_encontrado"),
                "carpeta_encontrada": r.get("carpeta_encontrada"),
                "extension"         : r.get("extension"),
                "confianza"         : r.get("confianza"),
                "patron_match"      : r.get("patron_match"),
                "observaciones"     : r.get("observaciones"),
                "auditado_en"       : ahora,
            }
            for r in resultados_v2
        ]

        for i in range(0, len(registros), 500):
            db.table("evidencias_auditoria").insert(registros[i : i + 500]).execute()

        print(f"      ☁️  {len(registros)} registros → tabla 'evidencias_auditoria' ✅")
    except Exception as e:
        err = str(e)
        if "PGRST205" in err or "evidencias_auditoria" in err:
            print(
                "      ⚠️  Tabla 'evidencias_auditoria' no existe en Supabase.\n"
                "          Esto NO afecta la auditoría ni la tabla 'verificaciones'.\n"
                "          Ejecuta sql/crear_evidencias_auditoria.sql para habilitarla."
            )
        elif "SSL" in err or "CERTIFICATE" in err.upper():
            print(
                "      ⚠️  Error SSL al guardar auditoría enriquecida.\n"
                "          Los resultados de Drive sí se procesaron; solo falla el guardado.\n"
                f"          Detalle: {err[:120]}"
            )
        else:
            print(f"      ⚠️  Error guardando auditoría enriquecida: {e}")


# ── CLASIFICACIÓN DE GUÍAS ───────────────────────────────────────────────────

def _clasificar_y_ordenar_guias(
    grupos_v2: dict[str, list[dict]],
) -> list[tuple[str, str]]:
    """
    Clasifica y ordena las guías para el menú y los reportes.

    Orden de salida:
      1. PRINCIPAL — primera guía de cada número (01, 02, 03...)
      2. CRUZADA   — guías adicionales que comparten número con una PRINCIPAL
      3. INDUCCIÓN — GUIA_00 siempre al final (base común, no opaca a las demás)

    Dentro de cada categoría el orden es alfabético.
    Retorna lista de (nombre_guia, etiqueta).
    """
    def _num(nombre: str) -> int:
        m = re.search(r'(?:GUIA|Gu[ií]a)_?0*(\d+)', nombre, re.IGNORECASE)
        return int(m.group(1)) if m else 50

    por_numero: dict[int, list[str]] = {}
    for g in grupos_v2:
        por_numero.setdefault(_num(g), []).append(g)

    principales, cruzadas, induccion = [], [], []
    for num in sorted(por_numero):
        grupo = sorted(por_numero[num])
        if num == 0:
            induccion.extend(grupo)
        else:
            for i, g in enumerate(grupo):
                (principales if i == 0 else cruzadas).append(g)

    return (
        [(g, "PRINCIPAL")  for g in principales]
        + [(g, "CRUZADA")  for g in cruzadas]
        + [(g, "INDUCCIÓN") for g in induccion]
    )


# ── FLUJO PRINCIPAL ───────────────────────────────────────────────────────────

def auditar_ficha(ruta_excel: str, service):
    nombre_excel = os.path.basename(ruta_excel)
    print(f"\n{'█'*60}")
    print(f"  AUDITORÍA DE FICHA: {nombre_excel}")
    print(f"  Esta ejecución procesa SOLO esta ficha y la guía seleccionada.")
    print(f"  No se tocan registros de otras fichas en Supabase.")
    print(f"{'█'*60}")

    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    except Exception as e:
        print(f"❌ No se pudo abrir el Excel: {e}")
        return

    hoja_bd = wb["BASE DE DATOS"] if "BASE DE DATOS" in wb.sheetnames else None
    if not hoja_bd:
        print("❌ No se encontró la hoja 'BASE DE DATOS'")
        return

    # Leer datos con conversión robusta
    programa_raw = _leer_celda_texto(hoja_bd["C3"])
    ficha        = _leer_celda_texto(hoja_bd["C4"])   # ← ahora convierte int a str
    colegio      = _leer_celda_texto(hoja_bd["C7"])

    # Extraer siglas del nombre del archivo (ej: CGS, CJLL, CRFK, CTSM)
    siglas_m = re.match(r'^([A-Z]+)', nombre_excel)
    siglas   = siglas_m.group(1) if siglas_m else "GRP"

    if not programa_raw:
        print("❌ No se encontró el programa en celda C3")
        return

    if not ficha:
        # Si C4 está vacío, extraer número del nombre del archivo
        nums = re.findall(r'\d{7}', nombre_excel)
        ficha = nums[0] if nums else nombre_excel

    print(f"   Ficha    : {ficha}")
    print(f"   Colegio  : {colegio}")
    print(f"   Siglas   : {siglas}")

    # Resolver programa: Supabase primero, fallback local segundo
    programa_por_ficha, fuente_programa = resolver_programa_ficha(ficha)
    if fuente_programa == "supabase":
        print(f"   Programa : {programa_raw}  →  {programa_por_ficha}  ☁️  (Supabase)")
    elif fuente_programa == "fallback_local":
        print(f"   Programa : {programa_raw}  →  {programa_por_ficha}  💾 (fallback local)")
    else:
        print(f"   Programa : {programa_raw}  (sin resolver — se intentará desde celda C3)")

    # ── Intentar cargar desde Supabase (Fase 2) ──────────────────────────────
    grupos_v2 = _cargar_actividades_por_guia(programa_raw, programa_canonical=programa_por_ficha)
    usar_v2   = bool(grupos_v2)

    # Universo completo para detección de EVIDENCIA_CRUZADA (carga única por ficha)
    if usar_v2:
        todas_actividades = cargar_actividades()
        print(f"   🌐 Universo: {len(todas_actividades)} actividades (detección cruzada activa)")

        # ── Clasificar, ordenar y mostrar debug de guías ─────────────────────
        clasificadas  = _clasificar_y_ordenar_guias(grupos_v2)
        etiqueta_guia = {g: e for g, e in clasificadas}

        g_principales = [g for g, e in clasificadas if e == "PRINCIPAL"]
        g_cruzadas    = [g for g, e in clasificadas if e == "CRUZADA"]
        g_induccion   = [g for g, e in clasificadas if e == "INDUCCIÓN"]

        n_guias      = len(grupos_v2)
        n_acts       = sum(len(v) for v in grupos_v2.values())
        prog_display = programa_por_ficha or programa_raw

        print(f"\n   {'─'*54}")
        print(f"   🗂  Programa  : {prog_display}")
        if g_principales:
            print(f"   🔎 Guía principal detectada  : {g_principales[0]}")
        if g_cruzadas:
            print(f"   🔎 Guías cruzadas detectadas : {g_cruzadas}")
        if g_induccion:
            print(f"   🔎 Inducción detectada       : {', '.join(g_induccion)}")
        print(f"   🔎 Orden final de guías      : {[g for g,_ in clasificadas]}")
        print(f"\n   📚 Guías ({n_guias}):")
        for g, etiqueta in clasificadas:
            print(f"      • [{etiqueta}] {g}  ({len(grupos_v2[g])} actividades)")
        print(f"   📝 Total     : {n_acts} actividades distribuidas en {n_guias} guía(s)")
        print(f"   {'─'*54}\n")

        # ── Selección de guía a auditar ───────────────────────────────────────
        guias_ordenadas = [g for g, _ in clasificadas]   # orden ya aplicado
        print(f"   Guías disponibles:")
        print(f"   0. Todas las guías ({len(guias_ordenadas)})")
        for i, (g, etiqueta) in enumerate(clasificadas, 1):
            print(f"   {i}. {g}  ({len(grupos_v2[g])} actividades)  [{etiqueta}]")

        sel_guia = input("\n   Seleccione la guía a auditar (número o 0 para todas): ").strip()

        if sel_guia == "0":
            print(f"\n   📚 Auditando todas las guías ({len(guias_ordenadas)})")
        elif sel_guia.isdigit() and 1 <= int(sel_guia) <= len(guias_ordenadas):
            guia_elegida = guias_ordenadas[int(sel_guia) - 1]
            grupos_v2    = {guia_elegida: grupos_v2[guia_elegida]}
            etiqueta_sel = etiqueta_guia.get(guia_elegida, "")
            print(f"\n   🎯 Guía seleccionada: {guia_elegida}  [{etiqueta_sel}]  ({len(grupos_v2[guia_elegida])} actividades)")
        else:
            print("❌ Selección inválida")
            return
    else:
        todas_actividades = []

    if not usar_v2:
        # Fallback: leer desde PDFs locales (comportamiento original)
        carpeta_guias = _resolver_carpeta_guias(programa_raw)
        if not carpeta_guias:
            print(f"❌ No se encontró carpeta de guías para: '{programa_raw}'")
            print(f"   Agrega una entrada en MAPA_PROGRAMAS o ejecuta sincronizador_tabla.py")
            return
        guias = _leer_evidencias_de_guias(carpeta_guias)
        if not guias:
            return
    else:
        # Para _generar_reporte_excel construimos un dict compatible:
        # { nombre_guia: [nombre_esperado, ...] }
        # grupos_v2 ya está filtrado a la selección del usuario
        guias = {
            guia_id: [a["nombre_esperado"] for a in actos if a.get("nombre_esperado")]
            for guia_id, actos in grupos_v2.items()
        }

    aprendices = _leer_aprendices(hoja_bd)
    print(f"\n   👥 Aprendices encontrados: {len(aprendices)}")

    print(f"\n   0. Auditar TODOS")
    for i, a in enumerate(aprendices, 1):
        print(f"   {i}. {a['nombre']}")

    sel = input("\n   ¿Qué aprendiz auditar? (número o 0 para todos): ").strip()

    if sel == "0":
        aprendices_a_auditar = aprendices
    elif sel.isdigit() and 1 <= int(sel) <= len(aprendices):
        aprendices_a_auditar = [aprendices[int(sel) - 1]]
        print(f"\n   🎯 Auditando solo: {aprendices_a_auditar[0]['nombre']}")
    else:
        print("❌ Selección inválida")
        return

    print(f"   🔍 Iniciando auditoría {'(Fase 2 — motor enriquecido)' if usar_v2 else '(modo PDF)'}...\n")
    resultados = []

    for aprendiz in aprendices_a_auditar:
        nombre = aprendiz["nombre"]
        link   = aprendiz["link"]
        print(f"   🧐 {nombre}")

        resultado_aprendiz = {
            "nombre"    : nombre,
            "link"      : link,
            "guias"     : {},
            "guias_v2"  : {},
            "guias_v2_adv": {},   # { guia: { actividad_id: texto_advertencia } } — solo FORMATO_INCORRECTO
        }

        if not link or "http" not in str(link):
            print(f"      ⚠️  Sin link de Drive")
            resultados.append(resultado_aprendiz)
            continue

        folder_id = _extraer_id_carpeta(link)
        if not folder_id:
            print(f"      ❌ No se pudo extraer el ID del link")
            resultados.append(resultado_aprendiz)
            continue

        if usar_v2:
            # Listar archivos UNA SOLA VEZ por aprendiz
            try:
                archivos_drive = listar_archivos_con_info_service(service, folder_id)
                print(f"      📂 {len(archivos_drive)} archivo(s) en Drive")
            except Exception as e:
                print(f"      ❌ Error listando Drive: {e}")
                resultados.append(resultado_aprendiz)
                continue

            todos_resultados_v2: list[dict] = []

            # Usar el orden clasificado; filtrar a las guías activas en grupos_v2
            _guias_loop   = [(g, grupos_v2[g]) for g, _ in clasificadas if g in grupos_v2]
            n_guias_total = len(_guias_loop)
            for idx_guia, (nombre_guia, actividades) in enumerate(_guias_loop, 1):
                try:
                    etiqueta_actual = etiqueta_guia.get(nombre_guia, "")
                    # Resolver subcarpeta específica para esta guía (si está definida)
                    archivos_guia = _resolver_archivos_para_guia(archivos_drive, nombre_guia)

                    # ── Trazabilidad por guía ─────────────────────────────────
                    n_arch    = len(archivos_guia)
                    n_totales = len(archivos_drive)
                    arch_info = (
                        f"{n_arch}/{n_totales} archivos  ← subfiltrado a subcarpeta"
                        if n_arch < n_totales
                        else f"{n_arch} archivos"
                    )
                    print(f"\n      ┌─ Guía [{idx_guia}/{n_guias_total}] [{etiqueta_actual}]: {nombre_guia}")
                    print(f"      │  Actividades esperadas : {len(actividades)}")
                    print(f"      │  Archivos en Drive     : {arch_info}")

                    # ── Pre-flight: diagnóstico de datos en Supabase ─────────
                    vacios = [a for a in actividades if not (a.get("nombre_esperado") or "").strip()]
                    pct_vacios = len(vacios) / len(actividades) * 100 if actividades else 0
                    if vacios:
                        nivel = "🚨" if pct_vacios >= 20 else "⚠️ "
                        print(f"      │  {nivel} {len(vacios)}/{len(actividades)} ({pct_vacios:.0f}%) actividades SIN nombre_esperado")
                        print(f"      │      → el motor usará actividad_resumen como fallback")
                        if pct_vacios >= 20:
                            print(f"      │      → MÁS DEL 20% VACÍO — revisa el Excel maestro")
                        print(f"      │      → fix: python sincronizador_tabla.py --guia \"{nombre_guia}\"")

                    # ── Tabla de actividades: actividad_id | patrón | resumen ─
                    print(f"      │  Patrones de búsqueda ({len(actividades)} actividades):")
                    print(f"      │  {'ID':<14}  {'nombre_esperado':<35}  actividad_resumen")
                    print(f"      │  {'─'*80}")
                    for a in actividades:
                        a_id   = (a.get("actividad_id") or "?")[:13]
                        n_esp  = (a.get("nombre_esperado") or "").strip()
                        n_res  = (a.get("actividad_resumen") or "")[:34]
                        n_vars = len(a.get("variantes_permitidas") or [])
                        src    = n_esp[:33] if n_esp else f"[fallback] {n_res[:24]}"
                        flag   = " ⚠️" if not n_esp else ("" if n_vars == 0 else f" +{n_vars}v")
                        print(f"      │  {a_id:<14}  {src:<35}  {n_res}{flag}")

                    # ── Log de archivos en Drive ─────────────────────────────
                    print(f"      │  Archivos encontrados ({len(archivos_guia)}):")
                    for f in archivos_guia[:12]:
                        mime_short = (f.get("mime_type") or "").replace("application/vnd.google-apps.", "gapps.")
                        ext_info   = f.get("extension") or mime_short[:20] or "?"
                        print(f"      │     '{f['nombre']}'  [{ext_info}]  /{f.get('folder_path','') or '(raíz)'}")
                    if len(archivos_guia) > 12:
                        print(f"      │     … y {len(archivos_guia)-12} archivo(s) más")
                    print(f"      └─ Auditando...")

                    # Activar debug detallado del motor cuando hay vacíos o resultado 0
                    _debug_motor = bool(vacios)
                    resultados_motor = MotorAuditoria(
                        actividades, archivos_guia, todas_actividades,
                        debug=_debug_motor,
                    ).auditar()
                    todos_resultados_v2.extend(resultados_motor)

                    # ── Resultado detallado por actividad ────────────────────
                    for r in resultados_motor:
                        estado = r["estado"]
                        a_id   = r.get("actividad_id", "?")
                        n_esp  = (r.get("nombre_esperado") or "")[:40]
                        if estado == EstadoAuditoria.OK.value:
                            print(f"         ✓ OK         [{a_id}] '{n_esp}' ← '{r.get('archivo_encontrado','')}' ({r.get('confianza')}%)")
                        elif estado == EstadoAuditoria.NOMBRE_DIFERENTE.value:
                            print(f"         ~ NOMBRE_DIF [{a_id}] '{n_esp}' ← '{r.get('archivo_encontrado','')}' ({r.get('confianza')}%)")
                        elif estado in (EstadoAuditoria.FALTA.value, EstadoAuditoria.OPCIONAL_FALTANTE.value):
                            print(f"         ✗ {estado:<16} [{a_id}] '{n_esp}' — sin candidato")
                        else:
                            print(f"         · {estado:<16} [{a_id}] '{n_esp}' ← '{r.get('archivo_encontrado','')}'")

                    # Bool para _guardar_en_supabase (tabla verificaciones, dashboard).
                    # Formato incorrecto sigue siendo ENTREGADO — el tipo es solo advertencia.
                    reporte_bool = {
                        r["nombre_esperado"]: es_entregada(r)
                        for r in resultados_motor
                        if r.get("nombre_esperado")
                    }
                    # Estado rico por actividad_id para _generar_reporte_excel_v2
                    reporte_v2 = {
                        r["actividad_id"]: r["estado"]
                        for r in resultados_motor
                        if r.get("actividad_id")
                    }
                    # Advertencias de formato: solo entradas FORMATO_INCORRECTO
                    # Alimentan los comentarios de celda en el Excel.
                    reporte_v2_adv = {
                        r["actividad_id"]: r.get("observaciones") or "formato diferente al esperado"
                        for r in resultados_motor
                        if r.get("actividad_id")
                        and r["estado"] == EstadoAuditoria.FORMATO_INCORRECTO.value
                    }
                    entregadas  = sum(reporte_bool.values())
                    ok_estricto = sum(
                        1 for r in resultados_motor
                        if r["estado"] == EstadoAuditoria.OK.value
                    )
                    con_adv     = sum(
                        1 for r in resultados_motor
                        if r["estado"] in (
                            EstadoAuditoria.NOMBRE_DIFERENTE.value,
                            EstadoAuditoria.CARPETA_INCORRECTA.value,
                            EstadoAuditoria.FORMATO_INCORRECTO.value,
                        )
                    )
                    total = len(actividades)
                    adv_str = f"  ({con_adv} con advertencia ~⚠)" if con_adv else ""
                    if ok_estricto == 0 and entregadas > 0:
                        print(f"      ⚠️  {nombre_guia[:40]}: {entregadas}/{total} entregadas{adv_str}")
                        print(f"           → Ninguna con nombre/carpeta exactos — revisa carpeta_drive")
                    else:
                        print(f"      ✅ {nombre_guia[:40]}: {entregadas}/{total}{adv_str}")
                    resultado_aprendiz["guias"][nombre_guia]        = reporte_bool
                    resultado_aprendiz["guias_v2"][nombre_guia]     = reporte_v2
                    resultado_aprendiz["guias_v2_adv"][nombre_guia] = reporte_v2_adv
                except Exception as e:
                    print(f"      ❌ Error en {nombre_guia}: {e}")

            # ── Resumen por aprendiz ─────────────────────────────────────
            from collections import Counter as _Counter
            _estados = _Counter(r["estado"] for r in todos_resultados_v2)
            _faltan  = [r["actividad_id"] for r in todos_resultados_v2
                        if r["estado"] == EstadoAuditoria.FALTA.value and r.get("actividad_id")]
            _iconos  = {"OK":"✓","NOMBRE_DIFERENTE":"~","CARPETA_INCORRECTA":"📁",
                        "EVIDENCIA_CRUZADA":"↔","FALTA":"✗","OPCIONAL_FALTANTE":"○"}
            print(f"\n      {'─'*50}")
            print(f"      Resumen: {nombre}")
            for est, cnt in sorted(_estados.items()):
                falta_ids = f"  [{', '.join(_faltan[:6])}{'…' if len(_faltan)>6 else ''}]" if est == EstadoAuditoria.FALTA.value else ""
                print(f"      {_iconos.get(est,'·')} {est:<22}: {cnt}{falta_ids}")
            print(f"      {'─'*50}")

            # Guardar resultados ricos en evidencias_auditoria
            _guardar_auditoria_enriquecida(
                todos_resultados_v2, ficha, colegio, siglas, nombre
            )

        else:
            # Flujo original con PDFs
            for nombre_guia, evidencias in guias.items():
                try:
                    reporte = verificar_evidencias_en_carpeta(
                        service, folder_id, evidencias, debug=False
                    )
                    entregadas = sum(reporte.values())
                    total      = len(evidencias)
                    print(f"      📂 {nombre_guia[:40]}: {entregadas}/{total}")
                    resultado_aprendiz["guias"][nombre_guia] = reporte
                except Exception as e:
                    print(f"      ❌ Error en {nombre_guia}: {e}")

        resultados.append(resultado_aprendiz)

    # Reporte Excel local
    fecha_hoy   = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_base = re.sub(r'\.xlsx$', '', nombre_excel, flags=re.IGNORECASE)
    ruta_salida = os.path.join(CARPETA_REPORTES, f"Reporte_{nombre_base}_{fecha_hoy}.xlsx")
    if usar_v2:
        _generar_reporte_excel_v2(
            ficha, programa_raw, resultados, ruta_salida,
            guias_a_incluir=sorted(grupos_v2.keys()),
        )
    else:
        _generar_reporte_excel(ficha, programa_raw, resultados, guias, ruta_salida)

    # Guardar en Supabase
    _guardar_en_supabase(resultados, ficha, colegio, siglas)

    print(f"\n✅ Ficha {ficha} — {siglas} completada.")


def ejecutar_todas_las_fichas():
    print("\n🤖 GUÍABOT — AUDITORÍA DE EVIDENCIAS")
    print(f"   Buscando fichas en: {CARPETA_LISTADOS}\n")

    excels = sorted(glob.glob(os.path.join(CARPETA_LISTADOS, "*.xlsx")))

    if not excels:
        print(f"❌ No se encontraron archivos Excel en {CARPETA_LISTADOS}")
        return

    print(f"   Fichas encontradas ({len(excels)}):")
    for i, e in enumerate(excels, 1):
        print(f"   {i}. {os.path.basename(e)}")

    print(f"\n   0. Auditar TODAS")
    seleccion = input("\n   ¿Qué ficha auditar? (número o 0 para todas): ").strip()

    if seleccion == "0":
        fichas_a_auditar = excels
    elif seleccion.isdigit() and 1 <= int(seleccion) <= len(excels):
        fichas_a_auditar = [excels[int(seleccion) - 1]]
    else:
        print("❌ Selección inválida")
        return

    service = conectar_drive()
    if not service:
        return

    for ruta_excel in fichas_a_auditar:
        auditar_ficha(ruta_excel, service)

    print(f"\n🎉 Proceso completado. Reportes en: {CARPETA_REPORTES}/")


# ── API: función no-interactiva llamada desde el endpoint REST ───────────────

def ejecutar_auditoria(
    ficha_id: str,
    guia_id: str | None = None,
    service=None,
) -> dict:
    """
    Versión no-interactiva de auditar_ficha para uso desde /api/auditar.

    - No usa input(); siempre audita todos los aprendices.
    - service: Drive service pre-construido (si None lo crea aquí).
    - Retorna dict con status, conteos y duración; nunca lanza excepción.
    """
    import time
    t0 = time.time()

    # ── Encontrar el Excel de la ficha ────────────────────────────────────────
    patron     = os.path.join(CARPETA_LISTADOS, f"*{ficha_id}*.xlsx")
    candidatos = [f for f in glob.glob(patron)
                  if not os.path.basename(f).startswith("~$")]
    if not candidatos:
        return {"status": "error", "mensaje": f"No se encontró Excel para ficha {ficha_id}"}
    ruta_excel   = candidatos[0]
    nombre_excel = os.path.basename(ruta_excel)

    # ── Drive ─────────────────────────────────────────────────────────────────
    if service is None:
        service = conectar_drive()
    if not service:
        return {"status": "error", "mensaje": "No se pudo conectar a Google Drive"}

    # ── Abrir Excel ───────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    except Exception as e:
        return {"status": "error", "mensaje": f"No se pudo abrir el Excel: {e}"}

    hoja_bd = wb.get("BASE DE DATOS")
    if not hoja_bd:
        return {"status": "error", "mensaje": "No se encontró la hoja 'BASE DE DATOS'"}

    programa_raw = _leer_celda_texto(hoja_bd["C3"])
    ficha        = _leer_celda_texto(hoja_bd["C4"]) or ficha_id
    colegio      = _leer_celda_texto(hoja_bd["C7"])
    siglas_m     = re.match(r'^([A-Z]+)', nombre_excel)
    siglas       = siglas_m.group(1) if siglas_m else "GRP"

    if not ficha:
        nums  = re.findall(r'\d{7}', nombre_excel)
        ficha = nums[0] if nums else ficha_id

    # ── Programa ──────────────────────────────────────────────────────────────
    programa_canonical, _ = resolver_programa_ficha(ficha)

    # ── Actividades ───────────────────────────────────────────────────────────
    grupos_v2 = _cargar_actividades_por_guia(programa_raw, programa_canonical=programa_canonical)
    if not grupos_v2:
        return {"status": "error", "mensaje": "No se encontraron actividades en Supabase"}

    # Filtrar a la guía indicada (coincidencia exacta o parcial)
    if guia_id:
        if guia_id not in grupos_v2:
            matches = [g for g in grupos_v2
                       if guia_id.lower() in g.lower() or g.lower() in guia_id.lower()]
            if not matches:
                return {
                    "status": "error",
                    "mensaje": (f"Guía '{guia_id}' no encontrada. "
                                f"Disponibles: {sorted(grupos_v2)[:10]}"),
                }
            guia_id = matches[0]
        grupos_v2 = {guia_id: grupos_v2[guia_id]}

    todas_actividades = cargar_actividades()
    clasificadas      = _clasificar_y_ordenar_guias(grupos_v2)
    etiqueta_guia     = {g: e for g, e in clasificadas}
    aprendices        = _leer_aprendices(hoja_bd)

    resultados:         list[dict] = []
    total_entregadas    = 0
    total_actividades   = 0
    total_advertencias  = 0

    for aprendiz in aprendices:
        nombre = aprendiz["nombre"]
        link   = aprendiz["link"]
        resultado_ap = {
            "nombre": nombre, "link": link,
            "guias": {}, "guias_v2": {}, "guias_v2_adv": {},
        }

        if not link or "http" not in str(link):
            resultados.append(resultado_ap)
            continue

        folder_id = _extraer_id_carpeta(link)
        if not folder_id:
            resultados.append(resultado_ap)
            continue

        try:
            archivos_drive = listar_archivos_con_info_service(service, folder_id)
        except Exception as e:
            print(f"   ⚠️  Drive error para {nombre}: {e}")
            resultados.append(resultado_ap)
            continue

        todos_v2: list[dict] = []
        _loop = [(g, grupos_v2[g]) for g, _ in clasificadas if g in grupos_v2]

        for nombre_guia, actividades in _loop:
            try:
                archivos_guia    = _resolver_archivos_para_guia(archivos_drive, nombre_guia)
                resultados_motor = MotorAuditoria(
                    actividades, archivos_guia, todas_actividades,
                ).auditar()
                todos_v2.extend(resultados_motor)

                reporte_bool = {
                    r["nombre_esperado"]: es_entregada(r)
                    for r in resultados_motor if r.get("nombre_esperado")
                }
                reporte_v2 = {
                    r["actividad_id"]: r["estado"]
                    for r in resultados_motor if r.get("actividad_id")
                }
                reporte_v2_adv = {
                    r["actividad_id"]: r.get("observaciones") or "formato diferente"
                    for r in resultados_motor
                    if r.get("actividad_id")
                    and r["estado"] == EstadoAuditoria.FORMATO_INCORRECTO.value
                }

                total_entregadas   += sum(reporte_bool.values())
                total_actividades  += len(actividades)
                total_advertencias += sum(
                    1 for r in resultados_motor
                    if r["estado"] in (
                        EstadoAuditoria.NOMBRE_DIFERENTE.value,
                        EstadoAuditoria.CARPETA_INCORRECTA.value,
                        EstadoAuditoria.FORMATO_INCORRECTO.value,
                    )
                )

                resultado_ap["guias"][nombre_guia]        = reporte_bool
                resultado_ap["guias_v2"][nombre_guia]     = reporte_v2
                resultado_ap["guias_v2_adv"][nombre_guia] = reporte_v2_adv
            except Exception as e:
                print(f"   ⚠️  Error en {nombre_guia} / {nombre}: {e}")

        _guardar_auditoria_enriquecida(todos_v2, ficha, colegio, siglas, nombre)
        resultados.append(resultado_ap)

    # ── Excel + Supabase ──────────────────────────────────────────────────────
    fecha_hoy   = datetime.now().strftime("%Y%m%d_%H%M")
    nombre_base = re.sub(r'\.xlsx$', '', nombre_excel, flags=re.IGNORECASE)
    ruta_salida = os.path.join(CARPETA_REPORTES, f"Reporte_{nombre_base}_{fecha_hoy}.xlsx")
    _generar_reporte_excel_v2(
        ficha, programa_raw, resultados, ruta_salida,
        guias_a_incluir=sorted(grupos_v2.keys()),
    )
    _guardar_en_supabase(resultados, ficha, colegio, siglas)

    return {
        "status"           : "ok",
        "ficha"            : ficha,
        "programa"         : programa_canonical or programa_raw,
        "guia"             : guia_id or "todas",
        "entregadas"       : total_entregadas,
        "total"            : total_actividades,
        "advertencias"     : total_advertencias,
        "aprendices"       : len([r for r in resultados if r.get("link")]),
        "duracion_segundos": round(time.time() - t0, 1),
    }


if __name__ == "__main__":
    ejecutar_todas_las_fichas()